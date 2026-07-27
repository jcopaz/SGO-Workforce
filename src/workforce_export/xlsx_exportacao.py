"""Exportacao XLSX (Incremento 11).

Abas conforme docs/14_EXPORTACOES.md: Resumo, HH por categoria, Jornadas,
Pausas, Falhas, Qualidade, Dicionario de dados. "HH por OS" nao existe
porque o sistema ainda nao modela OS (fora de escopo ate o Incremento 13).
"HH por ativo" e limitada aos atendimentos de falha, unico lugar onde
`ativo` existe hoje (DadosFalha.ativo, texto livre - Incremento 6).

Regra inegociavel aplicada (docs/27 secao 3.7): toda exportacao tem data
de geracao, periodo, filtros e usuario responsavel (aba Resumo) - e os
totais desta planilha vem exatamente de workforce_core.consolidacao, a
mesma fonte usada no painel (Incremento 9), para reconciliar com o
dashboard sem reimplementar o calculo.
"""

from __future__ import annotations

from datetime import timedelta
from pathlib import Path
from typing import List, Optional, Union

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from workforce_core.calculo import duracao_atividade_liquida, duracao_pausa
from workforce_core.catalogo import CatalogoMotivos, catalogo_padrao
from workforce_core.consolidacao import resumo_consolidado, taxa_qualidade_pulsos
from workforce_core.entities import Jornada, PulsoGps

from .csv_exportacao import linhas_falhas, linhas_jornadas
from .metadados import MetadadosExportacao


def _formatar_horas(duracao: timedelta) -> str:
    total_minutos = round(duracao.total_seconds() / 60)
    horas, minutos = divmod(total_minutos, 60)
    return f"{horas}h{minutos:02d}"


def _rotulo_categoria(categoria) -> str:
    return categoria.value if categoria is not None else "SEM_CATEGORIA"


def _escrever_tabela(aba: Worksheet, cabecalho: List[str], linhas: List[list]) -> None:
    aba.append(cabecalho)
    for linha in linhas:
        aba.append(linha)


def _aba_resumo(wb: Workbook, jornadas: List[Jornada], catalogo: CatalogoMotivos, metadados: MetadadosExportacao) -> None:
    aba = wb.active
    aba.title = "Resumo"
    resumo = resumo_consolidado(jornadas, catalogo)

    aba.append(["Campo", "Valor"])
    aba.append(["Usuario responsavel", metadados.usuario_responsavel])
    aba.append(["Data de geracao (tecnico, UTC)", metadados.data_geracao.isoformat()])
    aba.append(["Periodo inicio (tecnico)", metadados.periodo_inicio.isoformat() if metadados.periodo_inicio else ""])
    aba.append(["Periodo fim (tecnico)", metadados.periodo_fim.isoformat() if metadados.periodo_fim else ""])
    aba.append(["Filtros aplicados", str(metadados.filtros)])
    aba.append([])
    aba.append(["Quantidade de jornadas encerradas", resumo.quantidade_jornadas])
    aba.append(["HH bruto total (formatado)", _formatar_horas(resumo.jornada_bruta_total)])
    aba.append(["HH bruto total (segundos)", resumo.jornada_bruta_total.total_seconds()])
    aba.append(["HH classificado total (formatado)", _formatar_horas(resumo.tempo_classificado_total)])
    aba.append(["HH classificado total (segundos)", resumo.tempo_classificado_total.total_seconds()])
    aba.append(["HH nao classificado total (formatado)", _formatar_horas(resumo.tempo_nao_classificado_total)])
    aba.append(["HH nao classificado total (segundos)", resumo.tempo_nao_classificado_total.total_seconds()])


def _aba_hh_por_categoria(wb: Workbook, jornadas: List[Jornada], catalogo: CatalogoMotivos) -> None:
    aba = wb.create_sheet("HH por categoria")
    resumo = resumo_consolidado(jornadas, catalogo)
    linhas = [
        [_rotulo_categoria(categoria), _formatar_horas(duracao), duracao.total_seconds()]
        for categoria, duracao in sorted(
            resumo.por_categoria.items(), key=lambda kv: kv[1], reverse=True
        )
    ]
    _escrever_tabela(aba, ["Categoria", "HH (formatado)", "HH (segundos)"], linhas)


def _aba_hh_por_ativo(wb: Workbook, jornadas: List[Jornada]) -> None:
    aba = wb.create_sheet("HH por ativo")
    totais: dict[str, float] = {}
    for jornada in jornadas:
        for atividade in jornada.atividades:
            dados = atividade.dados_falha
            if dados is None or not dados.ativo or atividade.fim is None:
                continue
            totais[dados.ativo] = totais.get(dados.ativo, 0.0) + duracao_atividade_liquida(
                atividade
            ).total_seconds()

    linhas = [
        [ativo, _formatar_horas(timedelta(seconds=segundos)), segundos]
        for ativo, segundos in sorted(totais.items(), key=lambda kv: kv[1], reverse=True)
    ]
    _escrever_tabela(aba, ["Ativo", "HH (formatado)", "HH (segundos)"], linhas)


def _aba_jornadas(wb: Workbook, jornadas: List[Jornada]) -> None:
    aba = wb.create_sheet("Jornadas")
    cabecalho = [
        "jornada_id",
        "colaborador_matricula",
        "estado",
        "inicio",
        "fim",
        "jornada_bruta_segundos",
        "tempo_classificado_segundos",
        "tempo_nao_classificado_segundos",
    ]
    linhas = [[linha[campo] for campo in cabecalho] for linha in linhas_jornadas(jornadas)]
    _escrever_tabela(aba, cabecalho, linhas)


def _aba_pausas(wb: Workbook, jornadas: List[Jornada]) -> None:
    aba = wb.create_sheet("Pausas")
    cabecalho = ["jornada_id", "atividade_id", "motivo", "inicio", "fim", "duracao_segundos"]
    linhas = []
    for jornada in jornadas:
        for atividade in jornada.atividades:
            for pausa in atividade.pausas:
                linhas.append(
                    [
                        str(jornada.id),
                        str(atividade.id),
                        pausa.motivo,
                        pausa.inicio.isoformat() if pausa.inicio else "",
                        pausa.fim.isoformat() if pausa.fim else "",
                        duracao_pausa(pausa).total_seconds() if pausa.fim else "",
                    ]
                )
    _escrever_tabela(aba, cabecalho, linhas)


def _aba_falhas(wb: Workbook, jornadas: List[Jornada]) -> None:
    aba = wb.create_sheet("Falhas")
    cabecalho = [
        "jornada_id",
        "atividade_id",
        "inicio",
        "fim",
        "nota",
        "ativo",
        "sintoma",
        "objeto",
        "causa",
        "acao",
        "observacao",
        "completo",
    ]
    linhas = [[linha[campo] for campo in cabecalho] for linha in linhas_falhas(jornadas)]
    _escrever_tabela(aba, cabecalho, linhas)


def _aba_qualidade(wb: Workbook, pulsos: Optional[List[PulsoGps]]) -> None:
    aba = wb.create_sheet("Qualidade")
    aba.append(["Indicador", "Valor"])
    if pulsos is None:
        aba.append(["Pulsos GPS informados", "Nao (aba nao calculada)"])
        return
    taxa = taxa_qualidade_pulsos(pulsos)
    aba.append(["Quantidade de pulsos", len(pulsos)])
    aba.append(["Taxa de GPS valido (OK / avaliados)", taxa if taxa is not None else "N/A (nenhum avaliado)"])


_DICIONARIO_DADOS = [
    ("Resumo", "Usuario responsavel", "Quem gerou a exportacao (docs/27 secao 3.7)."),
    ("Resumo", "Filtros aplicados", "Filtros usados para selecionar as jornadas exportadas."),
    ("HH por categoria", "Categoria", "Categoria de docs/07_MOTOR_EVENTOS_E_HH.md; SEM_CATEGORIA quando o motivo nao tem categoria no catalogo."),
    ("HH por ativo", "Ativo", "Campo livre DadosFalha.ativo (Incremento 6); so existe para atendimentos de falha."),
    ("Jornadas", "estado", "NAO_INICIADA, ABERTA ou ENCERRADA."),
    ("Jornadas", "*_segundos", "Duracoes em segundos, para uso tecnico/recalculo."),
    ("Pausas", "motivo", "Codigo de motivo (ex.: PAUSA_TESTE) - catalogo oficial ainda pendente (ADR-0005)."),
    ("Falhas", "completo", "True se nota/ativo/sintoma/objeto/observacao estao todos preenchidos (docs/48_ADR_0021_ATENDIMENTO_DE_FALHA_CAMPO.md)."),
    ("Qualidade", "Taxa de GPS valido", "Proporcao de pulsos OK entre os avaliados (workforce_core.consolidacao.taxa_qualidade_pulsos)."),
]


def _aba_dicionario(wb: Workbook) -> None:
    aba = wb.create_sheet("Dicionario de dados")
    _escrever_tabela(aba, ["Aba", "Coluna", "Descricao"], [list(linha) for linha in _DICIONARIO_DADOS])


def exportar_xlsx(
    caminho: Union[str, Path],
    jornadas: List[Jornada],
    metadados: MetadadosExportacao,
    *,
    pulsos: Optional[List[PulsoGps]] = None,
    catalogo: Optional[CatalogoMotivos] = None,
) -> Path:
    catalogo = catalogo or catalogo_padrao()
    wb = Workbook()

    _aba_resumo(wb, jornadas, catalogo, metadados)
    _aba_hh_por_categoria(wb, jornadas, catalogo)
    _aba_hh_por_ativo(wb, jornadas)
    _aba_jornadas(wb, jornadas)
    _aba_pausas(wb, jornadas)
    _aba_falhas(wb, jornadas)
    _aba_qualidade(wb, pulsos)
    _aba_dicionario(wb)

    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    return caminho
