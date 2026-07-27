"""Testes do Incremento 11: exportacoes CSV, XLSX e GeoJSON.

Cobre a regra inegociavel de docs/27_ALINHAMENTO_OFICIAL_SGO_WORKFORCE_v1_2.md
secao 3.7: toda exportacao tem data de geracao, periodo, filtros e usuario
responsavel, e o total exportado reconcilia com o dashboard (mesma fonte
de calculo, workforce_core.consolidacao).
"""

import csv
import json
from datetime import datetime, timedelta
from uuid import uuid4

import openpyxl
import pytest

from workforce_core import MotorJornada, TipoEventoSecundario
from workforce_core.catalogo import catalogo_padrao
from workforce_core.consolidacao import resumo_consolidado
from workforce_core.entities import PulsoGps
from workforce_export import (
    MetadadosExportacao,
    exportar_csvs,
    exportar_geojson,
    exportar_xlsx,
    feature_collection_pontos,
    feature_collection_trajetorias,
    linhas_falhas,
    linhas_jornadas,
)


def _dt(hora, minuto, dia=1):
    return datetime(2026, 1, dia, hora, minuto)


def _jornada_completa(matricula="12345"):
    motor = MotorJornada(matricula)
    motor.iniciar_jornada(_dt(8, 0))
    motor.iniciar_evento_secundario(_dt(8, 0), TipoEventoSecundario.DESLOCAMENTO, "DESLOCAMENTO_TESTE")
    motor.encerrar_evento_secundario(_dt(8, 30))
    motor.iniciar_atividade(_dt(8, 30))
    motor.iniciar_pausa(_dt(9, 0), "PAUSA_TESTE")
    motor.finalizar_pausa(_dt(9, 10))
    motor.encerrar_atividade(_dt(11, 30))
    motor.iniciar_atendimento_falha(_dt(11, 30))
    motor.registrar_dados_falha(
        nota="1", ativo="ATIVO-X", sintoma="S", objeto="O", observacao="Obs"
    )
    motor.encerrar_atividade(_dt(12, 0))
    motor.encerrar_jornada(_dt(12, 0))
    return motor.jornada


def test_metadados_exige_usuario_responsavel():
    with pytest.raises(ValueError):
        MetadadosExportacao(usuario_responsavel="")


def test_metadados_sufixo_nome_arquivo_com_periodo():
    metadados = MetadadosExportacao(
        usuario_responsavel="tester",
        periodo_inicio=datetime(2026, 1, 1),
        periodo_fim=datetime(2026, 1, 31),
        data_geracao=datetime(2026, 2, 1, 10, 0, 0),
    )
    sufixo = metadados.sufixo_nome_arquivo()
    assert "20260101-20260131" in sufixo
    assert "20260201T100000Z" in sufixo


# ----------------------------------------------------------------------
# CSV
# ----------------------------------------------------------------------
def test_linhas_jornadas_reconcilia_com_consolidacao():
    jornada = _jornada_completa()
    linhas = linhas_jornadas([jornada])
    resumo = resumo_consolidado([jornada], catalogo_padrao())

    assert float(linhas[0]["jornada_bruta_segundos"]) == resumo.jornada_bruta_total.total_seconds()
    assert (
        float(linhas[0]["tempo_classificado_segundos"])
        == resumo.tempo_classificado_total.total_seconds()
    )


def test_linhas_falhas_marca_completo():
    jornada = _jornada_completa()
    linhas = linhas_falhas([jornada])
    assert len(linhas) == 1
    assert linhas[0]["completo"] is True
    assert linhas[0]["nota"] == "1"


def test_exportar_csvs_grava_arquivos_com_metadados(tmp_path):
    jornada = _jornada_completa()
    metadados = MetadadosExportacao(
        usuario_responsavel="tester",
        filtros={"colaborador": "12345"},
        periodo_inicio=_dt(0, 0),
        periodo_fim=_dt(23, 59),
    )

    caminhos = exportar_csvs(tmp_path, [jornada], [], metadados)

    nomes = {c.name for c in caminhos}
    assert any(n.startswith("jornadas_") for n in nomes)
    assert any(n.startswith("eventos_") for n in nomes)
    assert any(n.startswith("falhas_") for n in nomes)
    assert any(n.startswith("gps_") for n in nomes)
    assert any(n.startswith("metadados_") for n in nomes)

    caminho_metadados = next(c for c in caminhos if c.name.startswith("metadados_"))
    conteudo = json.loads(caminho_metadados.read_text(encoding="utf-8"))
    assert conteudo["usuario_responsavel"] == "tester"
    assert conteudo["filtros"] == {"colaborador": "12345"}

    caminho_eventos = next(c for c in caminhos if c.name.startswith("eventos_"))
    with open(caminho_eventos, encoding="utf-8-sig") as arquivo:
        linhas = list(csv.DictReader(arquivo))
    tipos = {linha["tipo_evento"] for linha in linhas}
    assert tipos == {"ATIVIDADE", "ATENDIMENTO_FALHA", "PAUSA", "DESLOCAMENTO"}


def test_exportar_csvs_com_pulsos(tmp_path):
    jornada = _jornada_completa()
    pulso = PulsoGps(
        jornada_id=jornada.id,
        colaborador_matricula="12345",
        latitude=-23.5,
        longitude=-46.6,
        precisao_metros=10,
        timestamp_dispositivo=_dt(8, 0),
    )
    metadados = MetadadosExportacao(usuario_responsavel="tester")

    caminhos = exportar_csvs(tmp_path, [jornada], [pulso], metadados)
    caminho_gps = next(c for c in caminhos if c.name.startswith("gps_"))
    with open(caminho_gps, encoding="utf-8-sig") as arquivo:
        linhas = list(csv.DictReader(arquivo))
    assert len(linhas) == 1
    assert linhas[0]["latitude"] == "-23.5"


# ----------------------------------------------------------------------
# XLSX
# ----------------------------------------------------------------------
def test_exportar_xlsx_reconcilia_totais_com_consolidacao(tmp_path):
    jornada = _jornada_completa()
    metadados = MetadadosExportacao(usuario_responsavel="tester")
    resumo = resumo_consolidado([jornada], catalogo_padrao())

    caminho = exportar_xlsx(tmp_path / "export.xlsx", [jornada], metadados)

    wb = openpyxl.load_workbook(caminho)
    assert set(wb.sheetnames) >= {
        "Resumo",
        "HH por categoria",
        "HH por ativo",
        "Jornadas",
        "Pausas",
        "Falhas",
        "Qualidade",
        "Dicionario de dados",
    }

    aba_resumo = wb["Resumo"]
    valores = {linha[0].value: linha[1].value for linha in aba_resumo.iter_rows() if linha[0].value}
    assert valores["Usuario responsavel"] == "tester"
    assert valores["HH bruto total (segundos)"] == resumo.jornada_bruta_total.total_seconds()


def test_exportar_xlsx_hh_por_ativo(tmp_path):
    jornada = _jornada_completa()
    metadados = MetadadosExportacao(usuario_responsavel="tester")

    caminho = exportar_xlsx(tmp_path / "export.xlsx", [jornada], metadados)
    wb = openpyxl.load_workbook(caminho)
    aba = wb["HH por ativo"]
    linhas = list(aba.iter_rows(values_only=True))
    assert linhas[0] == ("Ativo", "HH (formatado)", "HH (segundos)")
    assert linhas[1][0] == "ATIVO-X"


def test_exportar_xlsx_qualidade_sem_pulsos_nao_quebra(tmp_path):
    jornada = _jornada_completa()
    metadados = MetadadosExportacao(usuario_responsavel="tester")
    caminho = exportar_xlsx(tmp_path / "export.xlsx", [jornada], metadados)
    wb = openpyxl.load_workbook(caminho)
    assert wb["Qualidade"] is not None


# ----------------------------------------------------------------------
# GeoJSON
# ----------------------------------------------------------------------
def _pulso(jornada_id, segundos, lat=-23.5, lon=-46.6, matricula="12345"):
    return PulsoGps(
        jornada_id=jornada_id,
        colaborador_matricula=matricula,
        latitude=lat,
        longitude=lon,
        precisao_metros=10,
        timestamp_dispositivo=_dt(0, 0) + timedelta(seconds=segundos),
    )


def test_feature_collection_pontos_omite_matricula_por_padrao():
    jornada_id = uuid4()
    pulsos = [_pulso(jornada_id, 0)]

    colecao = feature_collection_pontos(pulsos)

    assert colecao["type"] == "FeatureCollection"
    propriedades = colecao["features"][0]["properties"]
    assert "colaborador_matricula" not in propriedades
    assert propriedades["jornada_id"] == str(jornada_id)


def test_feature_collection_pontos_inclui_matricula_se_pedido_explicitamente():
    jornada_id = uuid4()
    pulsos = [_pulso(jornada_id, 0, matricula="MAT-123")]

    colecao = feature_collection_pontos(pulsos, incluir_identificacao_pessoal=True)

    assert colecao["features"][0]["properties"]["colaborador_matricula"] == "MAT-123"


def test_feature_collection_pontos_coordenadas_lon_lat():
    jornada_id = uuid4()
    pulsos = [_pulso(jornada_id, 0, lat=-23.55, lon=-46.63)]
    colecao = feature_collection_pontos(pulsos)
    assert colecao["features"][0]["geometry"]["coordinates"] == [-46.63, -23.55]


def test_feature_collection_trajetorias_uma_por_jornada():
    jornada_1 = uuid4()
    jornada_2 = uuid4()
    pulsos_por_jornada = {
        jornada_1: [_pulso(jornada_1, i * 60, lat=-23.5 + i * 0.001) for i in range(5)],
        jornada_2: [_pulso(jornada_2, i * 60, lat=-22.9 + i * 0.001) for i in range(5)],
    }

    colecao = feature_collection_trajetorias(
        pulsos_por_jornada, distancia_simplificacao_metros=10
    )

    assert len(colecao["features"]) == 2
    for feature in colecao["features"]:
        assert feature["geometry"]["type"] == "LineString"
        assert "colaborador_matricula" not in feature["properties"]


def test_feature_collection_trajetorias_ignora_jornada_com_um_so_pulso():
    jornada_id = uuid4()
    pulsos_por_jornada = {jornada_id: [_pulso(jornada_id, 0)]}
    colecao = feature_collection_trajetorias(
        pulsos_por_jornada, distancia_simplificacao_metros=10
    )
    assert colecao["features"] == []


def test_exportar_geojson_grava_arquivo_valido(tmp_path):
    colecao = feature_collection_pontos([_pulso(uuid4(), 0)])
    caminho = exportar_geojson(tmp_path / "pontos.geojson", colecao)

    conteudo = json.loads(caminho.read_text(encoding="utf-8"))
    assert conteudo["type"] == "FeatureCollection"
    assert len(conteudo["features"]) == 1
